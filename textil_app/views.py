from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import login, logout, authenticate, get_user_model
from django.contrib.auth.forms import AuthenticationForm
from .models import Material, CustomUser, TikuvMashina, MaterialType, Color, Customer
from .forms import MaterialForm, CustomUserCreationForm, CustomAuthenticationForm, TikuvMashinaForm, ColorForm, MaterialTypeForm, CustomerForm
import datetime
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from django.views.decorators.http import require_POST
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from django.db.models import Q
from reportlab.platypus import Table, TableStyle
from reportlab.lib.units import mm
from rest_framework import generics
from .serializers import MaterialSerializer
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from django.utils.dateparse import parse_date
from django.db.models import Prefetch
from django.db import models
from datetime import timedelta
# from weasyprint import HTML  # WeasyPrint o'chirildi
from django.template.loader import render_to_string
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.db.models import Sum

User = get_user_model()

def is_admin(user):
    """Admin huquqlarini tekshirish"""
    return user.is_authenticated and (user.is_superuser or user.is_admin)

def login_view(request):
    """Login sahifasi"""
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        form = CustomAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f"Xush kelibsiz, {user.username}!")
                return redirect('home')
            else:
                messages.error(request, "Noto'g'ri foydalanuvchi nomi yoki parol!")
        else:
            messages.error(request, "Noto'g'ri foydalanuvchi nomi yoki parol!")
    else:
        form = CustomAuthenticationForm()
    
    return render(request, 'textil_app/login.html', {'form': form})

def logout_view(request):
    """Logout"""
    logout(request)
    messages.success(request, "Tizimdan chiqdingiz!")
    return redirect('login')

@login_required
def admin_panel(request):
    """Admin panel sahifasi"""
    if not is_admin(request.user):
        messages.error(request, "Sizda admin huquqlari yo'q!")
        return redirect('home')
    
    users = CustomUser.objects.all().order_by('-date_joined')
    materials = Material.objects.all().order_by('-kiritilgan_vaqt')[:10]
    tikuv_mashinalar = TikuvMashina.objects.all().order_by('raqami')[:5]
    
    context = {
        'users': users,
        'recent_materials': materials,
        'tikuv_mashinalar': tikuv_mashinalar,
        'total_users': users.count(),
        'total_materials': Material.objects.count(),
        'total_tikuv_mashinalar': TikuvMashina.objects.count(),
    }
    return render(request, 'textil_app/admin_panel.html', context)

# To'quv mashinalarini boshqarish
@login_required
@user_passes_test(is_admin)
def tikuv_mashina_list(request):
    """To'quv mashinalari ro'yxati"""
    mashinalar = TikuvMashina.objects.all().order_by('raqami')
    search = request.GET.get('search')
    holati = request.GET.get('holati')
    
    if holati in ['active', 'inactive', 'repair']:
        mashinalar = mashinalar.filter(holati=holati)
    if search:
        mashinalar = mashinalar.filter(
            Q(raqami__icontains=search) |
            Q(nomi__icontains=search)
        )
    
    context = {
        'mashinalar': mashinalar,
        'search': search,
        'holati': holati,
    }
    return render(request, 'textil_app/tikuv_mashina_list.html', context)

@login_required
@user_passes_test(is_admin)
def tikuv_mashina_create(request):
    """Yangi to'quv mashina qo'shish"""
    if request.method == 'POST':
        form = TikuvMashinaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "To'quv mashina muvaffaqiyatli qo'shildi!")
            return redirect('tikuv_mashina_list')
    else:
        form = TikuvMashinaForm()
    
    context = {
        'form': form,
        'title': 'Yangi To\'quv Mashina Qo\'shish'
    }
    return render(request, 'textil_app/tikuv_mashina_form.html', context)

@login_required
@user_passes_test(is_admin)
def tikuv_mashina_update(request, pk):
    """To'quv mashinani tahrirlash"""
    mashina = get_object_or_404(TikuvMashina, pk=pk)
    
    if request.method == 'POST':
        form = TikuvMashinaForm(request.POST, instance=mashina)
        if form.is_valid():
            form.save()
            messages.success(request, "To'quv mashina muvaffaqiyatli yangilandi!")
            return redirect('tikuv_mashina_list')
    else:
        form = TikuvMashinaForm(instance=mashina)
    
    context = {
        'form': form,
        'mashina': mashina,
        'title': 'To\'quv Mashinani Tahrirlash'
    }
    return render(request, 'textil_app/tikuv_mashina_form.html', context)

@login_required
@user_passes_test(is_admin)
def tikuv_mashina_delete(request, pk):
    """To'quv mashinani o'chirish"""
    mashina = get_object_or_404(TikuvMashina, pk=pk)
    
    # Mashinaga bog'langan materiallar borligini tekshirish
    materiallar = Material.objects.filter(tikuv_mashina=mashina)
    if materiallar.exists():
        messages.error(request, f"Bu mashinaga {materiallar.count()} ta material bog'langan. Avval materiallarni o'chiring!")
        return redirect('tikuv_mashina_list')
    
    if request.method == 'POST':
        mashina.delete()
        messages.success(request, "To'quv mashina muvaffaqiyatli o'chirildi!")
        return redirect('tikuv_mashina_list')
    
    context = {
        'mashina': mashina,
    }
    return render(request, 'textil_app/tikuv_mashina_confirm_delete.html', context)

@login_required
def home(request):
    """Bosh sahifa"""
    waiting_materials = Material.objects.filter(status='waiting').count()
    progress_materials = Material.objects.filter(status='progress').count()
    completed_materials = Material.objects.filter(status='completed').count()
    total_materials = Material.objects.count()
    recent_materials = Material.objects.all().order_by('-kiritilgan_vaqt')[:5]
    tikuv_mashinalar = TikuvMashina.objects.all().order_by('raqami')
    context = {
        'waiting_materials': waiting_materials,
        'progress_materials': progress_materials,
        'completed_materials': completed_materials,
        'total_materials': total_materials,
        'recent_materials': recent_materials,
        'tikuv_mashinalar': tikuv_mashinalar,
    }
    return render(request, 'textil_app/home.html', context)

@login_required
def material_list(request):
    """Materiallar ro'yxati"""
    materials = Material.objects.all().order_by('-kiritilgan_vaqt')
    search = request.GET.get('search')
    status = request.GET.get('status')
    if status in ['waiting', 'progress', 'completed']:
        materials = materials.filter(status=status)
    if search:
        materials = materials.filter(
            Q(partiya_raqami__icontains=search) |
            Q(buyurtmachi__icontains=search)
        )
    tikuv_mashinalar = TikuvMashina.objects.all().order_by('raqami')
    context = {
        'materials': materials,
        'search': search,
        'status': status,
        'tikuv_mashinalar': tikuv_mashinalar,
    }
    return render(request, 'textil_app/material_list.html', context)

@login_required
@user_passes_test(is_admin)
def material_create(request):
    """Yangi material qo'shish"""
    if request.method == 'POST':
        form = MaterialForm(request.POST)
        print(f"Form is_valid: {form.is_valid()}")  # Debug
        if form.is_valid():
            material = form.save(commit=False)
            material.created_by = request.user
            material.save()
            messages.success(request, "Material muvaffaqiyatli qo'shildi!")
            return redirect('material_list')
        else:
            print(f"Form errors: {form.errors}")  # Debug - form xatolarini ko'rsatish
            print(f"Form non_field_errors: {form.non_field_errors}")  # Debug
            messages.error(request, "Formada xatolik bor. Iltimos, barcha maydonlarni to'g'ri to'ldiring.")
    else:
        form = MaterialForm()
    
    context = {
        'form': form,
        'title': 'Yangi Material Qo\'shish'
    }
    return render(request, 'textil_app/material_form.html', context)

@login_required
def material_detail(request, pk):
    """Material haqida batafsil ma'lumot"""
    material = get_object_or_404(Material, pk=pk)
    context = {
        'material': material,
    }
    return render(request, 'textil_app/material_detail.html', context)

@login_required
@user_passes_test(is_admin)
def material_update(request, pk):
    """Materialni tahrirlash"""
    material = get_object_or_404(Material, pk=pk)
    
    if request.method == 'POST':
        form = MaterialForm(request.POST, instance=material)
        if form.is_valid():
            new_status = form.cleaned_data['status']
            if new_status == 'completed' and (material.tugatilgan_vaqt is None or material.status != 'completed'):
                material.tugatilgan_vaqt = timezone.now()
            elif new_status == 'waiting':
                material.tugatilgan_vaqt = None
            form.save()
            messages.success(request, "Material muvaffaqiyatli yangilandi!")
            return redirect('material_detail', pk=material.pk)
    else:
        form = MaterialForm(instance=material)
    
    context = {
        'form': form,
        'material': material,
        'title': 'Materialni Tahrirlash'
    }
    return render(request, 'textil_app/material_form.html', context)

@login_required
@user_passes_test(is_admin)
def material_delete(request, pk):
    """Materialni o'chirish"""
    material = get_object_or_404(Material, pk=pk)
    
    if request.method == 'POST':
        material.delete()
        messages.success(request, "Material muvaffaqiyatli o'chirildi!")
        return redirect('material_list')
    
    context = {
        'material': material,
    }
    return render(request, 'textil_app/material_confirm_delete.html', context)

@login_required
@user_passes_test(is_admin)
def export_materials_excel(request):
    period = request.GET.get('period', 'all')
    start = request.GET.get('start')
    end = request.GET.get('end')
    now = timezone.now()
    qs = Material.objects.all()
    # Sana oraliqlari bo'yicha filter
    if start:
        start_date = parse_date(start)
        if start_date:
            qs = qs.filter(kiritilgan_vaqt__date__gte=start_date)
    if end:
        end_date = parse_date(end)
        if end_date:
            qs = qs.filter(kiritilgan_vaqt__date__lte=end_date)
    # Oldingi period filterlari
    if period == 'day':
        qs = qs.filter(kiritilgan_vaqt__date=now.date())
    elif period == 'week':
        start_week = now - datetime.timedelta(days=now.weekday())
        qs = qs.filter(kiritilgan_vaqt__date__gte=start_week.date())
    elif period == 'month':
        qs = qs.filter(kiritilgan_vaqt__year=now.year, kiritilgan_vaqt__month=now.month)
    elif period == 'year':
        qs = qs.filter(kiritilgan_vaqt__year=now.year)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Materiallar"

    headers = [
        "ID", "Partiya raqami", "To'quv mashina", "Buyurtmachi", "Buyurtmachi tel", "Buyurtmachi email", "Material nomi",
        "Material rangi", "Material gramaji", "Kilogramm", "Ribana/Kashkor", "Bayka", "Tup/AEN", "Izoh", "Status", "Kiritilgan vaqt", "Tugatilgan vaqt"
    ]
    ws.append(headers)

    for m in qs:
        tikuv_mashina_nomi = m.tikuv_mashina.nomi if hasattr(m.tikuv_mashina, 'nomi') and m.tikuv_mashina.nomi else m.tikuv_mashina.raqami
        kiritilgan_vaqt = timezone.localtime(m.kiritilgan_vaqt).strftime('%d.%m.%Y %H:%M') if m.kiritilgan_vaqt else ''
        tugatilgan_vaqt = timezone.localtime(m.tugatilgan_vaqt).strftime('%d.%m.%Y %H:%M') if m.tugatilgan_vaqt else ''
        buyurtmachi_nomi = m.buyurtmachi.nomi if hasattr(m.buyurtmachi, 'nomi') else str(m.buyurtmachi)
        buyurtmachi_tel = getattr(m.buyurtmachi, 'telefon', None) or '—'
        buyurtmachi_email = getattr(m.buyurtmachi, 'email', None) or '—'
        material_turi = m.material_turi.nomi if hasattr(m.material_turi, 'nomi') else str(m.material_turi)
        material_rangi = m.material_rangi or '—'
        izoh = m.izoh or ''
        # Qo'shimcha maydonlar
        if hasattr(m, 'ribana_kashkor_turi') and m.ribana_kashkor_turi and hasattr(m, 'ribana_kashkor_kg') and m.ribana_kashkor_kg:
            ribana_kashkor = f"{m.get_ribana_kashkor_turi_display()} (kg): {m.ribana_kashkor_kg}"
        else:
            ribana_kashkor = "—"
        if hasattr(m, 'bayka_turi') and m.bayka_turi and hasattr(m, 'bayka_kg') and m.bayka_kg:
            bayka = f"{m.get_bayka_turi_display()} (kg): {m.bayka_kg}"
        else:
            bayka = "—"
        if hasattr(m, 'tup_aen_turi') and m.tup_aen_turi:
            tup_aen = m.get_tup_aen_turi_display()
        else:
            tup_aen = "—"
        ws.append([
            m.id,
            m.partiya_raqami,
            tikuv_mashina_nomi,
            buyurtmachi_nomi,
            buyurtmachi_tel,
            buyurtmachi_email,
            material_turi,
            material_rangi,
            m.material_gramaji_ko_rsatish,
            float(m.kilogramm),
            ribana_kashkor,
            bayka,
            tup_aen,
            izoh,
            m.get_status_display(),
            kiritilgan_vaqt,
            tugatilgan_vaqt
        ])

    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"materiallar_{period}_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

@login_required
@user_passes_test(is_admin)
@require_POST
def material_quick_status(request, pk):
    material = get_object_or_404(Material, pk=pk)
    new_status = request.POST.get('status')
    if new_status == 'completed':
        material.status = 'completed'
        material.tugatilgan_vaqt = timezone.now()
    elif new_status == 'waiting':
        material.status = 'waiting'
        material.tugatilgan_vaqt = None
    elif new_status == 'progress':
        material.status = 'progress'
        material.tugatilgan_vaqt = None
    material.save()
    messages.success(request, f"Material #{material.partiya_raqami} statusi '{material.get_status_display()}' ga o'zgartirildi!")
    
    # Qaysi sahifadan kelganini aniqlash
    next_url = request.POST.get('next', 'material_list')
    return redirect(next_url)

@login_required
@user_passes_test(is_admin)
def material_pdf(request, pk):
    material = get_object_or_404(Material, pk=pk)
    response = HttpResponse(content_type='application/pdf')
    filename = f"material_{material.partiya_raqami}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Sarlavha
    p.setFont("Helvetica-Bold", 18)
    from reportlab.lib import colors
    p.setFillColor(colors.HexColor("#1a237e"))
    p.drawCentredString(width / 2, height - 40, "Material Ma'lumotlari")
    p.setFillColor(colors.black)

    # Qo'shimcha maydonlar uchun qiymatlar
    if material.ribana_kashkor_turi and material.ribana_kashkor_kg:
        ribana_kashkor_str = f"{material.get_ribana_kashkor_turi_display()} (kg): {material.ribana_kashkor_kg}"
    else:
        ribana_kashkor_str = "—"
    if material.bayka_turi and material.bayka_kg:
        bayka_str = f"{material.get_bayka_turi_display()} (kg): {material.bayka_kg}"
    else:
        bayka_str = "—"
    if material.tup_aen_turi:
        tup_aen_str = material.get_tup_aen_turi_display()
    else:
        tup_aen_str = "—"
    customer_tel = material.buyurtmachi.telefon or "—"
    customer_email = material.buyurtmachi.email or "—"

    tikuv_mashina_nomi = material.tikuv_mashina.nomi if material.tikuv_mashina.nomi else material.tikuv_mashina.raqami
    kiritilgan_vaqt = timezone.localtime(material.kiritilgan_vaqt).strftime('%d.%m.%Y %H:%M') if material.kiritilgan_vaqt else ''
    tugatilgan_vaqt = timezone.localtime(material.tugatilgan_vaqt).strftime('%d.%m.%Y %H:%M') if material.tugatilgan_vaqt else 'Hali tugatilmagan'
    data = [
        ["Partiya raqami", material.partiya_raqami],
        ["To'quv mashina", tikuv_mashina_nomi],
        ["Buyurtmachi", material.buyurtmachi],
        ["Buyurtmachi tel", customer_tel],
        ["Buyurtmachi email", customer_email],
        ["Material nomi", material.material_turi],
        ["Material rangi", material.material_rangi],
        ["Material gramaji", material.material_gramaji_ko_rsatish],
        ["Kilogramm", f"{material.kilogramm} kg"],
        ["Ribana/Kashkor", ribana_kashkor_str],
        ["Bayka", bayka_str],
        ["Tup/AEN", tup_aen_str],
        ["Status", material.get_status_display()],
        ["Kiritilgan vaqt", kiritilgan_vaqt],
        ["Tugatilgan vaqt", tugatilgan_vaqt],
    ]

    # Jadvalni chizish
    table = Table(data, colWidths=[60*mm, 100*mm])
    table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 12),
        ('BACKGROUND', (0,0), (0,-1), colors.whitesmoke),
        ('TEXTCOLOR', (0,0), (0,-1), colors.HexColor("#1a237e")),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LINEBELOW', (0,0), (-1,-1), 0.25, colors.lightgrey),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))

    table.wrapOn(p, width, height)
    table.drawOn(p, 50, height - 420)

    # Mas'ul shaxs uchun joy
    y = height - 440 - len(data)*18
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, y, "Mas'ul shaxs:")
    p.line(130, y-2, 350, y-2)

    p.showPage()
    p.save()
    return response

class MaterialListCreateAPIView(generics.ListCreateAPIView):
    queryset = Material.objects.all().order_by('-kiritilgan_vaqt')
    serializer_class = MaterialSerializer

class MaterialRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Material.objects.all()
    serializer_class = MaterialSerializer

@api_view(['GET'])
@permission_classes([AllowAny])
def api_export_materials_excel(request):
    """
    Excel faylini yuklab olish uchun API endpoint.

    Tutorial:
    1. Endpoint: `/api/materials/export/`
    2. So'rov turi: GET
    3. Parametrlar (ixtiyoriy):
        - period=all|day|week|month|year
        - start=YYYY-MM-DD (boshlanish sanasi)
        - end=YYYY-MM-DD (tugash sanasi)
    4. Misol so'rovlar:
        - `/api/materials/export/?start=2024-06-01&end=2024-06-30`
        - `/api/materials/export/?period=week`
    5. Natija: Excel (xlsx) fayl sifatida yuklab olinadi.
    """
    return export_materials_excel(request)

@api_view(['GET'])
@permission_classes([AllowAny])
def api_material_pdf(request, pk):
    return material_pdf(request, pk)

@login_required
def mashina_zakazlar(request):
    mashinalar = TikuvMashina.objects.all().order_by('raqami')
    now = timezone.now()
    # Avval eng ko'p materialli mashinani topamiz (faqat waiting/progress)
    max_materials = 0
    for mashina in mashinalar:
        materials_count = Material.objects.filter(tikuv_mashina=mashina, status__in=['waiting', 'progress']).count()
        if materials_count > max_materials:
            max_materials = materials_count
    # Materiallarni matrix formatda tayyorlaymiz (faqat waiting/progress, eski yuqorida)
    materials_matrix = []
    for row in range(max_materials):
        row_materials = []
        for mashina in mashinalar:
            materials = Material.objects.filter(
                tikuv_mashina=mashina,
                status__in=['waiting', 'progress']
            ).order_by(
                models.Case(
                    models.When(status='progress', then=0),
                    models.When(status='waiting', then=1),
                    default=2,
                    output_field=models.IntegerField(),
                ),
                'kiritilgan_vaqt'
            )
            # Row indeksiga mos materialni olamiz
            if row < len(materials):
                mat = materials[row]
                is_late = False
                if mat.status == 'progress' and (now - mat.kiritilgan_vaqt) > timedelta(days=15):
                    is_late = True
                mat.is_late = is_late
                row_materials.append(mat)
            else:
                row_materials.append(None)
        materials_matrix.append(row_materials)
    # Pastdagi mashina statistikasi (faqat waiting/progress)
    mashina_stats = []
    for mashina in mashinalar:
        materials = Material.objects.filter(tikuv_mashina=mashina, status__in=['waiting', 'progress'])
        zakaz_soni = materials.count()
        umumiy_kg = materials.aggregate(total_kg=models.Sum('kilogramm'))['total_kg'] or 0
        
        mashina_stats.append({
            'raqami': mashina.raqami,
            'nomi': mashina.nomi,
            'zakaz_soni': zakaz_soni,
            'umumiy_kg': umumiy_kg,
        })
    context = {
        'mashinalar': mashinalar,
        'materials_matrix': materials_matrix,
        'max_materials': max_materials,
        'mashina_stats': mashina_stats,
    }
    return render(request, 'textil_app/mashina_zakazlar.html', context)

@login_required
@user_passes_test(is_admin)
def user_create(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f"Foydalanuvchi {user.username} muvaffaqiyatli qo'shildi!")
            return redirect('admin_panel')
    else:
        form = CustomUserCreationForm()
    context = {
        'form': form,
        'title': "Yangi Foydalanuvchi Qo'shish"
    }
    return render(request, 'textil_app/user_form.html', context)

@login_required
@user_passes_test(is_admin)
def user_update(request, pk):
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, f"Foydalanuvchi {user.username} muvaffaqiyatli yangilandi!")
            return redirect('admin_panel')
    else:
        form = CustomUserCreationForm(instance=user)
    context = {
        'form': form,
        'user': user,
        'title': "Foydalanuvchini Tahrirlash"
    }
    return render(request, 'textil_app/user_form.html', context)

@login_required
@user_passes_test(is_admin)
def user_delete(request, pk):
    user = get_object_or_404(User, pk=pk)
    if request.user == user:
        messages.error(request, "O'zingizni o'chira olmaysiz!")
        return redirect('admin_panel')
    if request.method == 'POST':
        username = user.username
        user.delete()
        messages.success(request, f"Foydalanuvchi {username} muvaffaqiyatli o'chirildi!")
        return redirect('admin_panel')
    context = {
        'user_to_delete': user,
    }
    return render(request, 'textil_app/user_confirm_delete.html', context)

@login_required
@user_passes_test(is_admin)
def user_list(request):
    users = CustomUser.objects.all().order_by('-date_joined')
    context = {
        'users': users,
    }
    return render(request, 'textil_app/user_list.html', context)

@login_required
@user_passes_test(is_admin)
def color_list(request):
    """Ranglar ro'yxati"""
    colors = Color.objects.all().order_by('-qo_shilgan_vaqt')
    search = request.GET.get('search')
    if search:
        colors = colors.filter(nomi__icontains=search)
    context = {
        'colors': colors,
        'search': search,
        'title': 'Ranglar ro\'yxati',
    }
    return render(request, 'textil_app/color_list.html', context)

@login_required
@user_passes_test(is_admin)
def color_create(request):
    """Yangi rang qo'shish"""
    if request.method == 'POST':
        form = ColorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Rang muvaffaqiyatli qo'shildi!")
            return redirect('color_list')
    else:
        form = ColorForm()
    context = {
        'form': form,
        'title': 'Yangi Rang Qo\'shish',
    }
    return render(request, 'textil_app/color_form.html', context)

@login_required
@user_passes_test(is_admin)
def color_update(request, pk):
    """Rangni tahrirlash"""
    color = get_object_or_404(Color, pk=pk)
    if request.method == 'POST':
        form = ColorForm(request.POST, instance=color)
        if form.is_valid():
            form.save()
            messages.success(request, "Rang muvaffaqiyatli yangilandi!")
            return redirect('color_list')
    else:
        form = ColorForm(instance=color)
    context = {
        'form': form,
        'color': color,
        'title': 'Rangni Tahrirlash',
    }
    return render(request, 'textil_app/color_form.html', context)

@login_required
@user_passes_test(is_admin)
def color_delete(request, pk):
    """Rangni o'chirish"""
    color = get_object_or_404(Color, pk=pk)
    if request.method == 'POST':
        color.delete()
        messages.success(request, "Rang muvaffaqiyatli o'chirildi!")
        return redirect('color_list')
    context = {
        'color': color,
        'title': 'Rangni O\'chirish',
    }
    return render(request, 'textil_app/color_confirm_delete.html', context)

@login_required
@user_passes_test(is_admin)
def materialtype_list(request):
    """Material turlari ro'yxati"""
    materialtypes = MaterialType.objects.all().order_by('-qo_shilgan_vaqt')
    search = request.GET.get('search')
    if search:
        materialtypes = materialtypes.filter(nomi__icontains=search)
    context = {
        'materialtypes': materialtypes,
        'search': search,
        'title': "Material nomlari ro'yxati",
    }
    return render(request, 'textil_app/materialtype_list.html', context)

@login_required
@user_passes_test(is_admin)
def materialtype_create(request):
    """Yangi material turi qo'shish"""
    if request.method == 'POST':
        form = MaterialTypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Material nomi muvaffaqiyatli qo'shildi!")
            return redirect('materialtype_list')
    else:
        form = MaterialTypeForm()
    context = {
        'form': form,
        'title': "Yangi Material Nomi Qo'shish",
    }
    return render(request, 'textil_app/materialtype_form.html', context)

@login_required
@user_passes_test(is_admin)
def materialtype_update(request, pk):
    """Material turini tahrirlash"""
    materialtype = get_object_or_404(MaterialType, pk=pk)
    if request.method == 'POST':
        form = MaterialTypeForm(request.POST, instance=materialtype)
        if form.is_valid():
            form.save()
            messages.success(request, "Material nomi muvaffaqiyatli yangilandi!")
            return redirect('materialtype_list')
    else:
        form = MaterialTypeForm(instance=materialtype)
    context = {
        'form': form,
        'materialtype': materialtype,
        'title': 'Material nomini tahrirlash',
    }
    return render(request, 'textil_app/materialtype_form.html', context)

@login_required
@user_passes_test(is_admin)
def materialtype_delete(request, pk):
    """Material turini o'chirish"""
    materialtype = get_object_or_404(MaterialType, pk=pk)
    if request.method == 'POST':
        materialtype.delete()
        messages.success(request, "Material nomi muvaffaqiyatli o'chirildi!")
        return redirect('materialtype_list')
    context = {
        'materialtype': materialtype,
        'title': "Material nomini o'chirish",
    }
    return render(request, 'textil_app/materialtype_confirm_delete.html', context)

@login_required
@user_passes_test(is_admin)
def customer_list(request):
    """Buyurtmachilar ro'yxati"""
    customers = Customer.objects.all().order_by('-qo_shilgan_vaqt')
    search = request.GET.get('search')
    if search:
        customers = customers.filter(nomi__icontains=search)
    context = {
        'customers': customers,
        'search': search,
        'title': "Buyurtmachilar ro'yxati",
    }
    return render(request, 'textil_app/customer_list.html', context)

@login_required
@user_passes_test(is_admin)
def customer_create(request):
    """Yangi buyurtmachi qo'shish"""
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Buyurtmachi muvaffaqiyatli qo'shildi!")
            return redirect('customer_list')
    else:
        form = CustomerForm()
    context = {
        'form': form,
        'title': "Yangi Buyurtmachi Qo'shish",
    }
    return render(request, 'textil_app/customer_form.html', context)

@login_required
@user_passes_test(is_admin)
def customer_update(request, pk):
    """Buyurtmachini tahrirlash"""
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, "Buyurtmachi muvaffaqiyatli yangilandi!")
            return redirect('customer_list')
    else:
        form = CustomerForm(instance=customer)
    context = {
        'form': form,
        'customer': customer,
        'title': 'Buyurtmachini tahrirlash',
    }
    return render(request, 'textil_app/customer_form.html', context)

@login_required
@user_passes_test(is_admin)
def customer_delete(request, pk):
    """Buyurtmachini o'chirish"""
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        customer.delete()
        messages.success(request, "Buyurtmachi muvaffaqiyatli o'chirildi!")
        return redirect('customer_list')
    context = {
        'customer': customer,
        'title': "Buyurtmachini o'chirish",
    }
    return render(request, 'textil_app/customer_confirm_delete.html', context)

@login_required
def mashina_zakazlar_pdf(request):
    mashinalar = TikuvMashina.objects.all().order_by('raqami')
    now = timezone.now()

    # Matrix uchun ma'lumotlar
    max_materials = 0
    mashina_materials = []
    jami_kg = []
    for mashina in mashinalar:
        materials = Material.objects.filter(
            tikuv_mashina=mashina,
            status__in=['waiting', 'progress', 'completed']
        ).order_by(
            models.Case(
                models.When(status='progress', then=0),
                models.When(status='waiting', then=1),
                models.When(status='completed', then=2),
                default=3,
                output_field=models.IntegerField(),
            ),
            'kiritilgan_vaqt'
        )
        mashina_materials.append(list(materials))
        jami_kg.append(sum([float(m.kilogramm) for m in materials]))
        if len(materials) > max_materials:
            max_materials = len(materials)

    # Jadval sarlavhasi (mashina nomlari)
    header = [m.nomi or m.raqami for m in mashinalar]

    # Matrix qatorlari
    matrix = []
    for row in range(max_materials):
        row_cells = []
        for col, materials in enumerate(mashina_materials):
            if row < len(materials):
                m = materials[row]
                # Katak ichidagi matn
                cell_text = f"""<b>#{m.partiya_raqami}</b>\n{m.buyurtmachi}\n{m.material_turi}, {m.material_rangi}\n{m.material_gramaji_ko_rsatish}, {m.kilogramm} kg\n{m.kiritilgan_vaqt.strftime('%d.%m.%Y')}\nStatus: {m.get_status_display()}"""
                row_cells.append(cell_text)
            else:
                row_cells.append("")
        matrix.append(row_cells)

    # Pastki qator (Итого)
    footer = [f"Итого: {kg} kg" if kg else "" for kg in jami_kg]

    # Jadval uchun data
    table_data = [header] + matrix + [footer]

    # PDF yaratish
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="zakazlar_matrix.pdf"'

    # Landscape format
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=10, leftMargin=10, topMargin=30, bottomMargin=20)
    elements = []
    styles = getSampleStyleSheet()
    elements.append(Paragraph("Mashinalar bo'yicha zakazlar", styles['Title']))
    elements.append(Spacer(1, 12))

    # Jadval
    col_width = 60 * mm
    table = Table(table_data, colWidths=[col_width]*len(header))

    # Jadval stillari
    style = TableStyle([
        # Sarlavha (yashil fon, oq matn)
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),

        # Pastki qator (footer)
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, -1), (-1, -1), 'CENTER'),

        # Chegaralar
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ])

    # Kataklar rangini statusga qarab o'zgartirish
    for row_idx, row in enumerate(matrix, start=1):
        for col_idx, cell in enumerate(row):
            if cell:
                m = mashina_materials[col_idx][row_idx-1]
                if m.status == 'progress':
                    style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.lightgreen)
                elif m.status == 'waiting':
                    style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.whitesmoke)
                elif m.status == 'completed':
                    style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.lightyellow)
                # Agar kechikkan bo'lsa (masalan, progress va 15 kundan ko'p bo'lsa)
                if m.status == 'progress' and (now - m.kiritilgan_vaqt).days > 15:
                    style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.red)

    table.setStyle(style)
    elements.append(table)
    doc.build(elements)
    return response

@login_required
@user_passes_test(is_admin)
def mashina_zakazlar_excel(request):
    mashinalar = TikuvMashina.objects.all().order_by('raqami')
    now = timezone.now()

    # Matrix uchun ma'lumotlar
    max_materials = 0
    mashina_materials = []
    jami_kg = []
    for mashina in mashinalar:
        materials = Material.objects.filter(
            tikuv_mashina=mashina,
            status__in=['waiting', 'progress', 'completed']
        ).order_by(
            models.Case(
                models.When(status='progress', then=0),
                models.When(status='waiting', then=1),
                models.When(status='completed', then=2),
                default=3,
                output_field=models.IntegerField(),
            ),
            'kiritilgan_vaqt'
        )
        mashina_materials.append(list(materials))
        jami_kg.append(sum([float(m.kilogramm) for m in materials]))
        if len(materials) > max_materials:
            max_materials = len(materials)

    # Excel workbook va sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mashina Zakazlar"

    # Sarlavha (header)
    header = [m.nomi or m.raqami for m in mashinalar]
    for col, name in enumerate(header, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill("solid", fgColor="228B22")  # Yashil
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Qatorlar (buyurtmalar)
    for row in range(max_materials):
        for col, materials in enumerate(mashina_materials):
            if row < len(materials):
                m = materials[row]
                status = m.get_status_display()
                sana = m.kiritilgan_vaqt.strftime('%d.%m.%Y')
                value = f"#{m.partiya_raqami}\n{m.buyurtmachi}\n{m.material_turi}, {m.material_rangi}\n{m.material_gramaji_ko_rsatish}, {m.kilogramm} kg\n{sana}\nStatus: {status}"
                cell = ws.cell(row=row+2, column=col+1, value=value)
                # Rang statusga qarab
                if m.status == 'progress':
                    cell.fill = PatternFill("solid", fgColor="90EE90")  # Yashil
                elif m.status == 'waiting':
                    cell.fill = PatternFill("solid", fgColor="FFFFFF")  # Oq
                elif m.status == 'completed':
                    cell.fill = PatternFill("solid", fgColor="F5F5F5")  # Kulrang
                if m.status == 'progress' and (now - m.kiritilgan_vaqt).days > 15:
                    cell.fill = PatternFill("solid", fgColor="FF6347")  # Qizil
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                cell.font = Font(size=10)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            else:
                cell = ws.cell(row=row+2, column=col+1, value="")
                cell.fill = PatternFill("solid", fgColor="F5F5F5")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Pastki qator (jami)
    for col, kg in enumerate(jami_kg, 1):
        cell = ws.cell(row=max_materials+2, column=col, value=f"Jami: {kg} kg" if kg else "")
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill("solid", fgColor="D3D3D3")  # Kulrang
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Ustun kengliklari
    for col in range(1, len(header)+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 32

    # Faylni yuborish
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"mashina_zakazlar_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

@login_required
@user_passes_test(is_admin)
@require_POST
def material_quick_mashina(request, pk):
    material = get_object_or_404(Material, pk=pk)
    mashina_id = request.POST.get('tikuv_mashina')
    if mashina_id:
        mashina = get_object_or_404(TikuvMashina, pk=mashina_id)
        material.tikuv_mashina = mashina
        material.save()
        messages.success(request, f"Material #{material.partiya_raqami} uchun to'quv mashina o'zgartirildi!")
    next_url = request.POST.get('next', 'material_list')
    return redirect(next_url)

@login_required
@user_passes_test(is_admin)
def mashina_statistika(request):
    now = timezone.now()
    week_ago = now - timedelta(days=7)
    mashinalar = TikuvMashina.objects.all()
    stats = []
    for mashina in mashinalar:
        materials = Material.objects.filter(
            tikuv_mashina=mashina,
            status='completed',
            tugatilgan_vaqt__gte=week_ago
        )
        count = materials.count()
        total_kg = materials.aggregate(Sum('kilogramm'))['kilogramm__sum'] or 0
        stats.append({
            'mashina': mashina.nomi or mashina.raqami,
            'count': count,
            'total_kg': float(total_kg),
        })
    context = {'stats': stats}
    return render(request, 'textil_app/mashina_statistika.html', context) 